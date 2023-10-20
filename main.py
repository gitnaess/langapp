# main.py

from website import create_app, db
from website.models import Question
import openpyxl

app = create_app()

def populate_db():
    # Delete all existing questions
    Question.query.delete()

    # Load your workbook and select your worksheet
    workbook = openpyxl.load_workbook('questions.xlsx')
    sheet = workbook.active
    
    # Go through each row in the worksheet and create a question object
    # Assumes that data starts in row 2 (to allow for header row)
    for row in range(2, sheet.max_row + 1):
        question_text = sheet.cell(row, 1).value
        option_a = sheet.cell(row, 2).value
        option_b = sheet.cell(row, 3).value
        option_c = sheet.cell(row, 4).value
        correct_option = sheet.cell(row, 5).value
        language = sheet.cell(row, 6).value
        qtype = sheet.cell(row, 7).value
        difficulty = sheet.cell(row, 8).value

        question = Question(
            question_text=question_text,
            option_a=option_a,
            option_b=option_b,
            option_c=option_c,
            correct_option=correct_option,
            language=language,
            qtype=qtype,
            difficulty=difficulty
        )
        db.session.add(question)
    
    # Commit all questions to the database
    db.session.commit()

if __name__ == '__main__':
    with app.app_context():
        # Only run once to avoid duplicating questions
        # Comment out after first run
        populate_db()
        
    app.run(debug=True)
